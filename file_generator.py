# file_generator.py
# Generates .xlsx data file, .xlsx metadata file, .zip archive, and latest/ copies

import os
import shutil
import zipfile
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import config

logger = logging.getLogger(__name__)


class OECDFileGenerator:
    """
    Generates all output files from parsed OECD data:
      1. OECD_VC_INV_DATA_{timestamp}.xlsx   — pivoted data (codes row, desc row, year rows)
      2. OECD_VC_INV_META_{timestamp}.xlsx   — one metadata row per column code
      3. OECD_VC_INV_{timestamp}.zip         — archive containing both xlsx files
      4. latest/ copies of all three files
    """

    def __init__(self):
        self.logger = logger

    # =========================================================================
    # DATA FILE
    # =========================================================================

    def create_data_file(self, parsed_result):
        """
        Write the pivoted data to an .xlsx file.

        Layout:
          Row 1  — column codes       (first cell blank = year column header)
          Row 2  — column descriptions (first cell blank)
          Row 3+ — year data rows     (first cell = year integer, then OBS_VALUE per column)

        Blank source values are written as empty cells (None).
        Numeric values are written as floats — full decimal precision preserved.

        Returns the output file path, or None on failure.
        """
        self.logger.info("Creating data .xlsx file...")

        os.makedirs(config.OUTPUT_DIR, exist_ok=True)

        filename  = config.DATA_FILE_PATTERN.format(timestamp=config.RUN_TIMESTAMP)
        file_path = os.path.join(config.OUTPUT_DIR, filename)

        column_codes = parsed_result['column_codes']
        column_descs = parsed_result['column_descs']
        pivot_rows   = parsed_result['pivot_rows']

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Data'

            # ----------------------------------------------------------------
            # Row 1: column codes
            # ----------------------------------------------------------------
            codes_row = [''] + column_codes          # first cell blank (year column)
            ws.append(codes_row)

            # ----------------------------------------------------------------
            # Row 2: column descriptions
            # ----------------------------------------------------------------
            descs_row = [''] + column_descs
            ws.append(descs_row)

            # ----------------------------------------------------------------
            # Rows 3+: year data
            # Write cell-by-cell so numeric raw strings are stored with
            # data_type='n', preserving the exact decimal from the source CSV.
            # (ws.append() would treat them as text strings.)
            # ----------------------------------------------------------------
            for xlsx_row, row_dict in enumerate(pivot_rows, start=3):
                # Column A: year (integer)
                ws.cell(row=xlsx_row, column=1, value=row_dict['year'])

                for col_idx, col_code in enumerate(column_codes, start=2):
                    value = row_dict.get(col_code)   # None or raw numeric string
                    if value is None:
                        pass   # leave cell empty
                    else:
                        cell = ws.cell(row=xlsx_row, column=col_idx)
                        cell.value     = value   # raw string from source CSV
                        cell.data_type = 'n'     # mark as number — no float truncation

            # ----------------------------------------------------------------
            # Column widths (auto-fit approximation)
            # ----------------------------------------------------------------
            ws.column_dimensions['A'].width = 10     # year column

            # ----------------------------------------------------------------
            # Save
            # ----------------------------------------------------------------
            wb.save(file_path)
            self.logger.info(
                f"Data file saved : {file_path} "
                f"({len(pivot_rows)} year rows × {len(column_codes)} columns)"
            )
            return file_path

        except Exception as e:
            self.logger.error(f"Error creating data file: {e}", exc_info=True)
            return None

    # =========================================================================
    # METADATA FILE
    # =========================================================================

    def create_meta_file(self, parsed_result):
        """
        Write one metadata row per column code to an .xlsx file.

        Each row contains:
          CODE, DESCRIPTION, FREQUENCY, MULTIPLIER, AGGREGATION_TYPE,
          UNIT_TYPE, DATA_TYPE, DATA_UNIT, SEASONALLY_ADJUSTED, ANNUALIZED,
          PROVIDER_MEASURE_URL, PROVIDER, SOURCE, SOURCE_DESCRIPTION,
          COUNTRY, DATASET

        COUNTRY is set to the individual country (Reference area label) for
        each column code so downstream systems can filter by country.

        Returns the output file path, or None on failure.
        """
        self.logger.info("Creating metadata .xlsx file...")

        os.makedirs(config.OUTPUT_DIR, exist_ok=True)

        filename  = config.META_FILE_PATTERN.format(timestamp=config.RUN_TIMESTAMP)
        file_path = os.path.join(config.OUTPUT_DIR, filename)

        column_codes = parsed_result['column_codes']
        column_descs = parsed_result['column_descs']
        country_map  = parsed_result['country_map']

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Metadata'

            # ----------------------------------------------------------------
            # Header row
            # ----------------------------------------------------------------
            ws.append(config.METADATA_COLUMNS)

            # ----------------------------------------------------------------
            # One metadata row per column code
            # ----------------------------------------------------------------
            defaults = config.METADATA_DEFAULTS

            for i, col_code in enumerate(column_codes):
                description = column_descs[i]
                country     = country_map.get(col_code, '')

                meta_row = {
                    'CODE':                 col_code,
                    'DESCRIPTION':          description,
                    'FREQUENCY':            defaults['FREQUENCY'],
                    'MULTIPLIER':           config.METADATA_MULTIPLIER,
                    'AGGREGATION_TYPE':     defaults['AGGREGATION_TYPE'],
                    'UNIT_TYPE':            defaults['UNIT_TYPE'],
                    'DATA_TYPE':            defaults['DATA_TYPE'],
                    'DATA_UNIT':            defaults['DATA_UNIT'],
                    'SEASONALLY_ADJUSTED':  defaults['SEASONALLY_ADJUSTED'],
                    'ANNUALIZED':           str(defaults['ANNUALIZED']),
                    'PROVIDER_MEASURE_URL': defaults['PROVIDER_MEASURE_URL'],
                    'PROVIDER':             defaults['PROVIDER'],
                    'SOURCE':               defaults['SOURCE'],
                    'SOURCE_DESCRIPTION':   defaults['SOURCE_DESCRIPTION'],
                    'COUNTRY':              country,
                    'DATASET':              defaults['DATASET'],
                }

                # Write row in exact METADATA_COLUMNS order
                ordered_row = [meta_row[col] for col in config.METADATA_COLUMNS]
                ws.append(ordered_row)

            wb.save(file_path)
            self.logger.info(
                f"Metadata file saved : {file_path} "
                f"({len(column_codes)} rows)"
            )
            return file_path

        except Exception as e:
            self.logger.error(f"Error creating metadata file: {e}", exc_info=True)
            return None

    # =========================================================================
    # ZIP ARCHIVE
    # =========================================================================

    def create_zip_file(self, data_file, meta_file):
        """
        Create a .zip archive containing the data and metadata xlsx files.
        Returns the zip file path, or None on failure.
        """
        self.logger.info("Creating zip archive...")

        zip_filename = config.ZIP_FILE_PATTERN.format(timestamp=config.RUN_TIMESTAMP)
        zip_path     = os.path.join(config.OUTPUT_DIR, zip_filename)

        files_to_zip = [f for f in [data_file, meta_file] if f and os.path.exists(f)]

        if not files_to_zip:
            self.logger.warning("No files to zip.")
            return None

        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for file_path in files_to_zip:
                    arcname = os.path.basename(file_path)
                    zf.write(file_path, arcname)
                    self.logger.info(f"  Added to zip: {arcname}")

            zip_size = os.path.getsize(zip_path)
            self.logger.info(
                f"Zip archive saved: {zip_path} ({zip_size:,} bytes)"
            )
            return zip_path

        except Exception as e:
            self.logger.error(f"Error creating zip file: {e}", exc_info=True)
            return None

    # =========================================================================
    # COPY TO LATEST
    # =========================================================================

    def copy_to_latest(self, file_paths):
        """
        Copy all generated files to the latest/ folder.
        Replaces the run timestamp with 'latest' in the filename.

        file_paths: dict with keys 'data', 'meta', 'zip'
        Returns dict of copied file paths.
        """
        self.logger.info("Copying files to 'latest' folder...")

        os.makedirs(config.LATEST_OUTPUT_DIR, exist_ok=True)

        copied = {}

        for file_type, source_path in file_paths.items():
            if not source_path or not os.path.exists(source_path):
                continue

            filename        = os.path.basename(source_path)
            latest_filename = filename.replace(config.RUN_TIMESTAMP, 'latest')
            dest_path       = os.path.join(config.LATEST_OUTPUT_DIR, latest_filename)

            try:
                shutil.copy2(source_path, dest_path)
                self.logger.info(f"  Copied {file_type}: {dest_path}")
                copied[file_type] = dest_path
            except Exception as e:
                self.logger.error(f"  Error copying {file_type}: {e}")

        return copied

    # =========================================================================
    # MAIN ENTRY POINT
    # =========================================================================

    def generate_files(self, parsed_result):
        """
        Generate all output files from the parsed result.

        Returns dict with all file paths, or None if data file creation fails.
        """
        print(f"\n{'='*60}")
        print("Generating output files...")
        print(f"{'='*60}\n")

        # ----------------------------------------------------------------
        # [1/4] Data xlsx
        # ----------------------------------------------------------------
        print("[1/4] Creating data xlsx file...")
        data_file = self.create_data_file(parsed_result)

        if not data_file:
            self.logger.error("Failed to create data file — aborting.")
            return None

        print(f"  [OK] {os.path.basename(data_file)}\n")

        # ----------------------------------------------------------------
        # [2/4] Metadata xlsx
        # ----------------------------------------------------------------
        print("[2/4] Creating metadata xlsx file...")
        meta_file = self.create_meta_file(parsed_result)

        if meta_file:
            print(f"  [OK] {os.path.basename(meta_file)}\n")
        else:
            print(f"  [FAILED] Metadata file\n")

        # ----------------------------------------------------------------
        # [3/4] Zip archive
        # ----------------------------------------------------------------
        print("[3/4] Creating zip archive...")
        zip_file = self.create_zip_file(data_file, meta_file)

        if zip_file:
            print(f"  [OK] {os.path.basename(zip_file)}\n")
        else:
            print(f"  [FAILED] Zip archive\n")

        # ----------------------------------------------------------------
        # [4/4] Copy to latest/
        # ----------------------------------------------------------------
        print("[4/4] Copying to 'latest' folder...")
        file_paths = {
            'data': data_file,
            'meta': meta_file,
            'zip':  zip_file,
        }
        copied = self.copy_to_latest(file_paths)
        print(f"  [OK] {len(copied)} file(s) copied\n")

        print(f"{'='*60}")
        print("File generation complete")
        print(f"{'='*60}\n")

        return {
            'data_file':        data_file,
            'meta_file':        meta_file,
            'zip_file':         zip_file,
            'latest_data_file': copied.get('data'),
            'latest_meta_file': copied.get('meta'),
            'latest_zip_file':  copied.get('zip'),
            'output_dir':       config.OUTPUT_DIR,
            'latest_dir':       config.LATEST_OUTPUT_DIR,
        }


def main():
    """Test file generator with sample parsed data (for development)."""
    import sys
    import logging

    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    # Minimal mock parsed result for standalone testing
    mock_result = {
        'source_file':  'test.csv',
        'column_codes': [
            'AUS.VCINV.VCT.USDV.A',
            'AUS.VCINV.SEED.USDV.A',
            'FRA.VCINV.VCT.USDV.A',
        ],
        'column_descs': [
            'Australia: Venture capital investments, Total, USD, current prices',
            'Australia: Venture capital investments, Seed, USD, current prices',
            'France: Venture capital investments, Total, USD, current prices',
        ],
        'country_map': {
            'AUS.VCINV.VCT.USDV.A':  'Australia',
            'AUS.VCINV.SEED.USDV.A': 'Australia',
            'FRA.VCINV.VCT.USDV.A':  'France',
        },
        'stage_map': {
            'AUS.VCINV.VCT.USDV.A':  'Total',
            'AUS.VCINV.SEED.USDV.A': 'Seed',
            'FRA.VCINV.VCT.USDV.A':  'Total',
        },
        'ref_area_map': {
            'AUS.VCINV.VCT.USDV.A':  'AUS',
            'AUS.VCINV.SEED.USDV.A': 'AUS',
            'FRA.VCINV.VCT.USDV.A':  'FRA',
        },
        'years': [2007, 2008, 2009],
        'pivot_rows': [
            {'year': 2007, 'AUS.VCINV.VCT.USDV.A': 439.676334938608,  'AUS.VCINV.SEED.USDV.A': None, 'FRA.VCINV.VCT.USDV.A': 1234.56},
            {'year': 2008, 'AUS.VCINV.VCT.USDV.A': 512.123456789012,  'AUS.VCINV.SEED.USDV.A': 10.5,  'FRA.VCINV.VCT.USDV.A': None},
            {'year': 2009, 'AUS.VCINV.VCT.USDV.A': None,              'AUS.VCINV.SEED.USDV.A': None,  'FRA.VCINV.VCT.USDV.A': 900.0},
        ],
        'row_count': 3,
        'col_count': 3,
    }

    generator = OECDFileGenerator()
    result = generator.generate_files(mock_result)

    if result:
        print("\nGenerated files:")
        print(f"  Data : {result['data_file']}")
        print(f"  Meta : {result['meta_file']}")
        print(f"  Zip  : {result['zip_file']}")


if __name__ == '__main__':
    main()
